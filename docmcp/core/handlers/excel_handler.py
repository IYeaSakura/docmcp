"""
Excel文档处理器模块

支持 .xls 和 .xlsx 格式的Excel文档处理。
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, BinaryIO, Dict, Iterator, List, Optional, Tuple, Union

import pandas as pd

from .base import BaseDocument, BaseDocumentHandler, BatchProcessor
from ..document import DocumentMetadata, DocumentType, ExtractedContent
from ..utils import TempFileManager, detect_file_type, chunk_file_reader

# 配置日志
logger = logging.getLogger(__name__)

# 尝试导入依赖库
try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl 未安装，XLSX功能受限")

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logger.warning("xlrd 未安装，XLS功能受限")

try:
    import xlwt
    XLWT_AVAILABLE = True
except ImportError:
    XLWT_AVAILABLE = False
    logger.warning("xlwt 未安装，XLS写入功能受限")


class ExcelDocument(BaseDocument):
    """
    Excel文档对象
    
    支持 .xls 和 .xlsx 格式的Excel文档。
    """
    
    # 大文件阈值 (10MB for Excel)
    LARGE_FILE_THRESHOLD = 10 * 1024 * 1024
    
    def __init__(
        self,
        file_path: Optional[Union[str, Path]] = None,
        file_stream: Optional[BinaryIO] = None,
        document_type: Optional[DocumentType] = None
    ):
        """
        初始化Excel文档对象
        
        Args:
            file_path: 文档文件路径
            file_stream: 文档文件流
            document_type: 文档类型
        """
        super().__init__(file_path, file_stream, document_type)
        self._workbook: Optional[Any] = None
        self._pandas_excel: Optional[pd.ExcelFile] = None
        self._temp_manager = TempFileManager()
        self._sheet_names: List[str] = []
    
    def load(self, **kwargs) -> "ExcelDocument":
        """
        加载Excel文档
        
        Args:
            read_only: 是否以只读模式打开（大文件推荐）
            data_only: 是否只读取数据（不读取公式）
            
        Returns:
            self: 支持链式调用
        """
        if self._is_loaded:
            return self
        
        read_only = kwargs.get('read_only', self.is_large_file())
        data_only = kwargs.get('data_only', True)
        
        try:
            # 确定文档类型
            if self._document_type == DocumentType.UNKNOWN:
                if self._file_path:
                    self._document_type = detect_file_type(self._file_path)
                elif self._file_stream:
                    self._document_type = detect_file_type(self._file_stream)
            
            # 加载文档
            if self._document_type == DocumentType.XLSX:
                self._load_xlsx(read_only=read_only, data_only=data_only)
            elif self._document_type == DocumentType.XLS:
                self._load_xls()
            else:
                self._try_auto_load()
            
            self._is_loaded = True
            self._update_metadata_from_file()
            
        except Exception as e:
            logger.error(f"加载Excel文档失败: {e}")
            raise ValueError(f"无法加载Excel文档: {e}")
        
        return self
    
    def _load_xlsx(self, read_only: bool = False, data_only: bool = True) -> None:
        """加载XLSX格式文档"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("需要安装 openpyxl 库: pip install openpyxl")
        
        if self._file_stream:
            self._file_stream.seek(0)
            self._workbook = load_workbook(
                self._file_stream,
                read_only=read_only,
                data_only=data_only
            )
        elif self._file_path:
            self._workbook = load_workbook(
                self._file_path,
                read_only=read_only,
                data_only=data_only
            )
        else:
            raise ValueError("没有可用的文件源")
        
        self._sheet_names = self._workbook.sheetnames
    
    def _load_xls(self) -> None:
        """加载XLS格式文档"""
        if not XLRD_AVAILABLE:
            raise ImportError("需要安装 xlrd 库: pip install xlrd")
        
        # xlrd 2.0+ 只支持xlsx，需要特殊处理
        try:
            if self._file_stream:
                self._file_stream.seek(0)
                self._workbook = xlrd.open_workbook(
                    file_contents=self._file_stream.read()
                )
            elif self._file_path:
                self._workbook = xlrd.open_workbook(str(self._file_path))
            else:
                raise ValueError("没有可用的文件源")
            
            self._sheet_names = self._workbook.sheet_names()
            
        except xlrd.biffh.XLRDError as e:
            # 可能是新版xlrd不支持xls，尝试使用pandas
            logger.warning(f"xlrd加载失败，尝试使用pandas: {e}")
            self._load_with_pandas()
    
    def _load_with_pandas(self) -> None:
        """使用pandas加载Excel"""
        if self._file_stream:
            self._file_stream.seek(0)
            self._pandas_excel = pd.ExcelFile(self._file_stream)
        elif self._file_path:
            self._pandas_excel = pd.ExcelFile(self._file_path)
        else:
            raise ValueError("没有可用的文件源")
        
        self._sheet_names = self._pandas_excel.sheet_names
    
    def _try_auto_load(self) -> None:
        """尝试自动检测并加载"""
        errors = []
        
        # 尝试XLSX
        try:
            self._load_xlsx()
            self._document_type = DocumentType.XLSX
            return
        except Exception as e:
            errors.append(f"XLSX: {e}")
        
        # 尝试XLS
        try:
            self._load_xls()
            self._document_type = DocumentType.XLS
            return
        except Exception as e:
            errors.append(f"XLS: {e}")
        
        # 尝试pandas
        try:
            self._load_with_pandas()
            self._document_type = DocumentType.XLSX
            return
        except Exception as e:
            errors.append(f"Pandas: {e}")
        
        raise ValueError(f"无法识别Excel格式: {'; '.join(errors)}")
    
    def get_sheet_names(self) -> List[str]:
        """
        获取所有工作表名称
        
        Returns:
            工作表名称列表
        """
        if not self._is_loaded:
            self.load()
        return self._sheet_names
    
    def extract_text(self, **kwargs) -> str:
        """
        提取纯文本内容
        
        Args:
            sheet_name: 指定工作表名称（可选）
            include_headers: 是否包含表头
            
        Returns:
            文档的纯文本内容
        """
        if not self._is_loaded:
            self.load()
        
        sheet_name = kwargs.get('sheet_name')
        include_headers = kwargs.get('include_headers', True)
        
        texts = []
        
        # 获取要处理的工作表
        sheet_names = [sheet_name] if sheet_name else self._sheet_names
        
        for name in sheet_names:
            sheet_data = self._get_sheet_data(name)
            
            texts.append(f"=== {name} ===")
            
            for row_idx, row in enumerate(sheet_data):
                if row_idx == 0 and not include_headers:
                    continue
                
                row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                if row_text.strip():
                    texts.append(row_text)
            
            texts.append("")  # 空行分隔
        
        return '\n'.join(texts)
    
    def _get_sheet_data(self, sheet_name: str) -> List[List[Any]]:
        """
        获取工作表数据
        
        Args:
            sheet_name: 工作表名称
            
        Returns:
            工作表数据（二维列表）
        """
        data = []
        
        if self._workbook:
            if OPENPYXL_AVAILABLE and isinstance(self._workbook, openpyxl.Workbook):
                sheet = self._workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    data.append(list(row))
            elif XLRD_AVAILABLE and isinstance(self._workbook, xlrd.Book):
                sheet = self._workbook.sheet_by_name(sheet_name)
                for row_idx in range(sheet.nrows):
                    row_data = []
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        row_data.append(cell.value)
                    data.append(row_data)
        
        elif self._pandas_excel:
            df = pd.read_excel(self._pandas_excel, sheet_name=sheet_name, header=None)
            data = df.values.tolist()
        
        return data
    
    def extract_content(self, **kwargs) -> ExtractedContent:
        """
        提取完整内容
        
        Args:
            sheet_name: 指定工作表名称（可选）
            
        Returns:
            包含所有工作表数据的完整内容
        """
        if not self._is_loaded:
            self.load()
        
        sheet_name = kwargs.get('sheet_name')
        
        content = ExtractedContent()
        
        # 获取要处理的工作表
        sheet_names = [sheet_name] if sheet_name else self._sheet_names
        
        for name in sheet_names:
            sheet_data = self._get_sheet_data(name)
            
            # 添加到sheets字典
            content.sheets[name] = {
                'data': sheet_data,
                'row_count': len(sheet_data),
                'column_count': len(sheet_data[0]) if sheet_data else 0
            }
            
            # 同时添加到tables列表
            if sheet_data:
                content.tables.append(sheet_data)
            
            # 构建文本
            for row in sheet_data:
                row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                if row_text.strip():
                    content.paragraphs.append(row_text)
        
        content.text = '\n'.join(content.paragraphs)
        
        return content
    
    def extract_metadata(self, **kwargs) -> DocumentMetadata:
        """
        提取文档元数据
        
        Returns:
            文档元数据对象
        """
        if not self._is_loaded:
            self.load()
        
        metadata = DocumentMetadata()
        
        # 从文件系统获取基本元数据
        self._update_metadata_from_file()
        metadata.file_size = self._metadata.file_size
        metadata.file_path = self._metadata.file_path
        metadata.file_name = self._metadata.file_name
        metadata.file_extension = self._metadata.file_extension
        
        # 添加工作表信息
        metadata.custom_properties['sheet_count'] = len(self._sheet_names)
        metadata.custom_properties['sheet_names'] = self._sheet_names
        
        # 尝试从openpyxl获取更多元数据
        if OPENPYXL_AVAILABLE and isinstance(self._workbook, openpyxl.Workbook):
            try:
                props = self._workbook.properties
                metadata.title = props.title
                metadata.subject = props.subject
                metadata.author = props.creator
                metadata.keywords = props.keywords
                metadata.comments = props.description
                
                if props.created:
                    metadata.created = props.created
                if props.modified:
                    metadata.modified = props.modified
                    
            except Exception as e:
                logger.debug(f"提取元数据失败: {e}")
        
        return metadata
    
    def to_dataframe(
        self,
        sheet_name: Optional[str] = None,
        header: Union[int, List[int], None] = 0,
        **kwargs
    ) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        转换为pandas DataFrame
        
        Args:
            sheet_name: 工作表名称（None表示所有工作表）
            header: 表头行号
            **kwargs: 传递给pd.read_excel的其他参数
            
        Returns:
            DataFrame或DataFrame字典
        """
        if not self._is_loaded:
            self.load()
        
        if self._pandas_excel:
            excel_file = self._pandas_excel
        elif self._file_path:
            excel_file = pd.ExcelFile(self._file_path)
        elif self._file_stream:
            self._file_stream.seek(0)
            excel_file = pd.ExcelFile(self._file_stream)
        else:
            raise ValueError("没有可用的文件源")
        
        if sheet_name is None:
            # 读取所有工作表
            result = {}
            for name in self._sheet_names:
                result[name] = pd.read_excel(
                    excel_file,
                    sheet_name=name,
                    header=header,
                    **kwargs
                )
            return result
        else:
            return pd.read_excel(
                excel_file,
                sheet_name=sheet_name,
                header=header,
                **kwargs
            )
    
    def save(self, output_path: Union[str, Path], **kwargs) -> "ExcelDocument":
        """
        保存文档
        
        Args:
            output_path: 输出文件路径
            
        Returns:
            self: 支持链式调用
        """
        if not self._is_loaded:
            self.load()
        
        output_path = Path(output_path)
        
        if self._workbook and OPENPYXL_AVAILABLE:
            if isinstance(self._workbook, openpyxl.Workbook):
                self._workbook.save(output_path)
            else:
                # 需要转换格式
                self._convert_and_save(output_path)
        else:
            raise ValueError("没有可保存的工作簿")
        
        return self
    
    def _convert_and_save(self, output_path: Path) -> None:
        """转换格式并保存"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("需要安装 openpyxl 库")
        
        # 创建新的工作簿
        new_wb = Workbook()
        
        # 复制数据
        for sheet_name in self._sheet_names:
            sheet_data = self._get_sheet_data(sheet_name)
            
            if sheet_name == self._sheet_names[0]:
                ws = new_wb.active
                ws.title = sheet_name
            else:
                ws = new_wb.create_sheet(title=sheet_name)
            
            for row_idx, row in enumerate(sheet_data, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        new_wb.save(output_path)
    
    def convert_to(
        self,
        target_type: DocumentType,
        output_path: Optional[Union[str, Path]] = None,
        **kwargs
    ) -> Union[str, Path, bytes]:
        """
        转换文档格式
        
        Args:
            target_type: 目标文档类型
            output_path: 输出路径（可选）
            
        Returns:
            转换后的文件路径或字节数据
        """
        if not self._is_loaded:
            self.load()
        
        if target_type == DocumentType.CSV:
            return self._convert_to_csv(output_path, **kwargs)
        elif target_type == DocumentType.XLSX:
            if output_path:
                self.save(output_path)
                return output_path
            else:
                buffer = io.BytesIO()
                if self._workbook and isinstance(self._workbook, openpyxl.Workbook):
                    self._workbook.save(buffer)
                return buffer.getvalue()
        elif target_type == DocumentType.XLS:
            return self._convert_to_xls(output_path, **kwargs)
        elif target_type == DocumentType.TXT:
            text = self.extract_text()
            if output_path:
                Path(output_path).write_text(text, encoding='utf-8')
                return output_path
            return text.encode('utf-8')
        else:
            raise ValueError(f"不支持的目标格式: {target_type}")
    
    def _convert_to_csv(
        self,
        output_path: Optional[Union[str, Path]] = None,
        sheet_name: Optional[str] = None,
        **kwargs
    ) -> Union[str, Path, Dict[str, Union[str, Path]]]:
        """
        转换为CSV
        
        Args:
            output_path: 输出路径
            sheet_name: 指定工作表（None表示所有）
            
        Returns:
            CSV文件路径或路径字典
        """
        if sheet_name is None:
            # 转换所有工作表
            results = {}
            for name in self._sheet_names:
                if output_path:
                    base_path = Path(output_path)
                    sheet_output = base_path.parent / f"{base_path.stem}_{name}.csv"
                else:
                    sheet_output = None
                
                results[name] = self._convert_to_csv(
                    sheet_output,
                    sheet_name=name,
                    **kwargs
                )
            
            return results
        
        # 转换单个工作表
        df = self.to_dataframe(sheet_name=sheet_name, header=0)
        
        if output_path:
            output_path = Path(output_path)
            df.to_csv(output_path, index=False, encoding='utf-8-sig')
            return output_path
        else:
            buffer = io.StringIO()
            df.to_csv(buffer, index=False, encoding='utf-8')
            return buffer.getvalue().encode('utf-8')
    
    def _convert_to_xls(
        self,
        output_path: Optional[Union[str, Path]] = None,
        **kwargs
    ) -> Union[str, Path]:
        """转换为XLS格式"""
        if not XLWT_AVAILABLE:
            raise ImportError("需要安装 xlwt 库: pip install xlwt")
        
        if not output_path:
            output_path = self._temp_manager.create_temp_file(suffix='.xls')
        else:
            output_path = Path(output_path)
        
        workbook = xlwt.Workbook()
        
        for sheet_name in self._sheet_names:
            sheet_data = self._get_sheet_data(sheet_name)
            
            ws = workbook.add_sheet(sheet_name)
            
            for row_idx, row in enumerate(sheet_data):
                for col_idx, value in enumerate(row):
                    ws.write(row_idx, col_idx, value)
        
        workbook.save(output_path)
        return output_path
    
    def close(self) -> None:
        """关闭文档，释放资源"""
        self._workbook = None
        self._pandas_excel = None
        self._temp_manager.cleanup()
        super().close()


class ExcelHandler(BaseDocumentHandler):
    """
    Excel文档处理器
    
    处理 .xls 和 .xlsx 格式的Excel文档。
    """
    
    @property
    def supported_types(self) -> List[DocumentType]:
        """返回支持的文档类型列表"""
        return [DocumentType.XLS, DocumentType.XLSX]
    
    @property
    def supported_extensions(self) -> List[str]:
        """返回支持的文件扩展名列表"""
        return ['.xls', '.xlsx']
    
    def create_document(
        self,
        file_path: Optional[Union[str, Path]] = None,
        file_stream: Optional[BinaryIO] = None
    ) -> ExcelDocument:
        """
        创建Excel文档对象
        
        Args:
            file_path: 文件路径
            file_stream: 文件流
            
        Returns:
            Excel文档对象
        """
        if file_path:
            self.validate_file(file_path)
        
        return ExcelDocument(file_path=file_path, file_stream=file_stream)
    
    def create_new_document(self) -> ExcelDocument:
        """
        创建新的空白Excel文档
        
        Returns:
            新的Excel文档对象
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("需要安装 openpyxl 库")
        
        doc = ExcelDocument()
        doc._workbook = Workbook()
        doc._sheet_names = ['Sheet']
        doc._document_type = DocumentType.XLSX
        doc._is_loaded = True
        return doc


# 便捷函数
def read_excel_to_dataframe(
    file_path: Union[str, Path],
    sheet_name: Optional[str] = None,
    **kwargs
) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
    """
    读取Excel到DataFrame
    
    Args:
        file_path: Excel文件路径
        sheet_name: 工作表名称
        **kwargs: 其他参数
        
    Returns:
        DataFrame或DataFrame字典
    """
    handler = ExcelHandler()
    doc = handler.create_document(file_path=file_path)
    with doc:
        return doc.to_dataframe(sheet_name=sheet_name, **kwargs)


def extract_tables_from_excel(file_path: Union[str, Path]) -> Dict[str, List[List[Any]]]:
    """
    从Excel提取所有表格
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        工作表名称到表格数据的字典
    """
    handler = ExcelHandler()
    doc = handler.create_document(file_path=file_path)
    with doc:
        content = doc.extract_content()
        return content.sheets
